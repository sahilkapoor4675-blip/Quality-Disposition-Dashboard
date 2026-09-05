import openpyxl, sqlite3, os

SRC = "/mnt/user-data/uploads/Quality_Disposition_Dashboard_White_Original_Layout_Fixed_KPI_ColorMatched.xlsm"
DB  = "/home/claude/qdash/quality.db"

if os.path.exists(DB):
    os.remove(DB)

wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
ws = wb["Disposition Data"]

conn = sqlite3.connect(DB)
cur = conn.cursor()
cur.execute("""
CREATE TABLE disposition (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    heat_no TEXT,
    work_center TEXT,
    grade TEXT,
    output_weight REAL,
    main_defect TEXT,
    defect_intensity TEXT,
    quality_decision TEXT,
    month TEXT,
    week TEXT,
    quarter TEXT,
    financial_year TEXT
)
""")

rows_to_insert = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    heat_no = row[2]   # C
    if heat_no is None or str(heat_no).strip() == "":
        continue
    work_center = row[5]        # F
    grade = row[6]               # G
    output_weight = row[7]       # H
    main_defect = row[14]        # O
    defect_intensity = row[15]   # P
    quality_decision = row[18]   # S
    month = row[22]              # W  _SourceMonth
    week = row[23]               # X
    quarter = row[24]            # Y
    fy = row[25]                 # Z

    def norm(v):
        if v is None:
            return ""
        return str(v).strip()

    rows_to_insert.append((
        norm(heat_no), norm(work_center), norm(grade),
        float(output_weight) if output_weight not in (None, "") else 0.0,
        norm(main_defect), norm(defect_intensity), norm(quality_decision),
        norm(month), norm(week), norm(quarter), norm(fy)
    ))

cur.executemany("""
INSERT INTO disposition
(heat_no, work_center, grade, output_weight, main_defect, defect_intensity,
 quality_decision, month, week, quarter, financial_year)
VALUES (?,?,?,?,?,?,?,?,?,?,?)
""", rows_to_insert)

# indexes for fast filtering
for col in ["work_center","grade","quality_decision","month","week","quarter","financial_year","defect_intensity","main_defect"]:
    cur.execute(f"CREATE INDEX idx_{col} ON disposition({col})")

conn.commit()

# Validation
cur.execute("SELECT COUNT(*) FROM disposition")
total = cur.fetchone()[0]
cur.execute("SELECT COUNT(*) FROM disposition WHERE output_weight IS NOT NULL AND output_weight <> 0")
with_weight = cur.fetchone()[0]
cur.execute("SELECT COUNT(*) FROM disposition WHERE output_weight >= 0")
weight_present = cur.fetchone()[0]

print("Total records inserted :", total)
print("Records with weight_present (>=0, i.e. not null) :", weight_present)
print("Records with non-zero output weight :", with_weight)

conn.close()
